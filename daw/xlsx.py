#!/usr/bin/python3

import openpyxl

def load(fn):
    return openpyxl.load_workbook(fn)

def findcell(sheet, needle):
    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            v = sheet.cell(r,c).value
            if type(v) == type(needle):
                if v==needle:
                    return r,c
    return None

def gettable(sheet, r0, c0):
    header = []
    table = {}
    for c in range(c0, sheet.max_column+1):
        v = sheet.cell(r0, c).value
        if v is None:
            break
        else:
            header.append(v)
            table[v] = []
    for r in range(r0+1, sheet.max_row+1):
        if sheet.cell(r, c0).value is None:
            break
        for dc in range(len(header)):
            v = sheet.cell(r, c0+dc).value
            table[header[dc]].append(v)
    return table
